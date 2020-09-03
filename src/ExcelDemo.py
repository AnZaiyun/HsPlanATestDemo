import xlrd
import xlwt
import os
import tkinter.filedialog
import logging
import openpyxl

def writeXls(filename, XlsData):
    logging.debug("开始写入数据")
    dirname = os.path.dirname(filename)
    logging.debug("xls文件写入目录：" + dirname)
    filename = filename.replace(dirname+"/","").replace(".xls","_tmp.xls")

    wb = xlwt.Workbook(dirname+"/"+filename)
    ws = wb.add_sheet('demo01', cell_overwrite_ok=True)  # 增加sheet
    for row in range(len(XlsData)):
        for col in range(len(XlsData[row])):
            ws.write(row,col,XlsData[row][col])

    wb.save(dirname+"/"+filename)

def readXls(filename):
    data = xlrd.open_workbook(filename)  # 打开Excel文件读取数据
    sheet_name = data.sheet_names()  # 获取所有sheet名称
    logging.debug("当前xls文件所有sheet名称："+str (sheet_name))

    sheet1 = data.sheet_by_index(0)  # 根据sheet索引或者名称获取sheet内容，同时获取sheet名称、列数、行数
    logging.debug('sheet1名称:{}  sheet1列数: {}  sheet1行数: {}'.format(sheet1.name, sheet1.ncols, sheet1.nrows))

    # 获取指定单元格的内容
    # print(sheet1.cell(1, 0).value)

    XlsData = []
    for i in range(sheet1.nrows):
        XlsData.append(sheet1.row_values(i))

    return XlsData

def wrXlsx(filename):
    dirname = os.path.dirname(filename)
    copyfilename = filename.replace(dirname+"/","").replace(".xlsx","_tmp.xlsx")
    logging.debug("开始读取数据（openpyxl）")
    wb = openpyxl.load_workbook(filename)  # 打开已有文件
    sheet_name = wb.sheetnames
    logging.debug("当前xlsx文件所有sheet名称：" + str(sheet_name))
    sheet1 = wb[sheet_name[0]]

    XlsxData = []
    for row in sheet1.rows:
        tmp = []
        for cell in row:
            tmp.append(cell.value)
        XlsxData.append(tmp)
    # print(XlsxData)

    logging.debug("开始复制数据（openpyxl）")

    wbr = openpyxl.Workbook(dirname+"/"+copyfilename)
    ws1 = wbr.create_sheet(sheet_name[0])  # 新建一个工作表，插入到最后

    for i in XlsxData:
        ws1.append(i)
    wbr.save(dirname+"/"+copyfilename)

# 设置显示日志级别
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')
filename = tkinter.filedialog.askopenfilename(filetypes=[('xlsx','.xlsx'),('xls','.xls')])
logging.debug("当前选择的程序文件为："+filename)

if filename.endswith(".xls"):
    # xlrd和xlwt也能操作xlsx格式的文件，不过在写入时只能写入xls格式，否则文件会打不开
    logging.debug("开始复制xls(2003)格式文件")
    XlsData = readXls(filename)  # 读取文件数据
    writeXls(filename,XlsData)  # 写入数据
elif filename.endswith(".xlsx"):
    logging.debug("开始复制xlsx(2007)格式文件")
    wrXlsx(filename)

